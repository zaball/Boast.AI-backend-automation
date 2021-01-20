import json
import jsonpath
import requests
import openpyxl


class Common:

    def __init__(self, file_name_path, sheet_name):
        global excel_file
        global excel_sheet
        excel_file = openpyxl.load_workbook(file_name_path)
        excel_sheet = excel_file[sheet_name]

    def fetch_row_count(self):
        rows = excel_sheet.max_row
        return rows

    def fetch_column_count(self):
        col = excel_sheet.max_column
        return col

    def fetch_key_names(self):
        columns = excel_sheet.max_column
        li = []
        for i in range (1, columns+1):
            cell = excel_sheet.cell(row=1, column=i)
            li.insert(i-1, cell.value)
        return li

    def update_request_with_data(self, row_number, json_request, key_list):
        columns = excel_sheet.max_column
        for i in range(1, columns+1):
            cell =excel_sheet.cell(row=row_number, column=i)
            json_request[key_list[i-1]] = cell.value
        return json_request

