import pytest
import json
import requests
import openpyxl
from boast_api_test.tests.data_practice import credentials


# simple way of adding multiple data
@pytest.mark.multiple_tenant_data01
def test_create_multiple_tenants():
    # API
    file = open('/home/adi/Boast/boast_api_test/tests/data_driven_testing/create_multiple_tenants.json')
    json_file = json.loads(file.read())
    # Excel Code
    excel_file = openpyxl.load_workbook('/home/adi/Boast/boast_api_test/tests/data_driven_testing/multiple_tenants.xlsx')
    excel_sheet = excel_file['Sheet1']
    rows = excel_sheet.max_row

    for i in range(2, rows+1):
        cell_status_name = excel_sheet.cell(row=i, column=1)
        cell_admin_name = excel_sheet.cell(row=i, column=2)
        cell_name = excel_sheet.cell(row=i, column=3)
        cell_db_name = excel_sheet.cell(row=i, column=4)
        cell_domain = excel_sheet.cell(row=i, column=5)
        cell_provisioned = excel_sheet.cell(row=i, column=6)
        cell_status = excel_sheet.cell(row=i, column=7)
        json_file['status_name'] = cell_status_name.value
        json_file['admin_name'] = cell_admin_name.value
        json_file['name'] = cell_name.value
        json_file['db_name'] = cell_db_name.value
        json_file['domain'] = cell_domain.value
        json_file['provisioned'] = cell_provisioned.value
        json_file['status'] = cell_status.value
        response = requests.post('http://badger.boast.net:8080/api/tenants/tenant/', json=json_file,
                                 auth=credentials)
        assert response.status_code == 201
        # pdb.set_trace()
