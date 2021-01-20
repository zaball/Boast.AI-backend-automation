import requests
import json
import pytest
import pdb
import openpyxl
from boast_api_test.tests.data_practice import credentials


# simple way of adding multiple data
@pytest.mark.multiple_user_data01
def test_create_multiple_users():
    # API
    file = open('/home/adi/Boast/boast_api_test/tests/data_driven_testing/create_multiple_users.json')
    json_file = json.loads(file.read())
    # Excel code
    excel_file = openpyxl.load_workbook('/home/adi/Boast/boast_api_test/tests/data_driven_testing/multiple_users.xlsx')
    excel_sheet = excel_file['Sheet1']
    rows = excel_sheet.max_row

    for i in range(2, rows+1):
        cell_first_name = excel_sheet.cell(row=i, column=1)
        cell_last_name = excel_sheet.cell(row=i, column=2)
        cell_username = excel_sheet.cell(row=i, column=3)
        cell_email = excel_sheet.cell(row=i, column=4)
        cell_is_active = excel_sheet.cell(row=i, column=5)
        cell_nick_name = excel_sheet.cell(row=i, column=6)
        cell_display_name = excel_sheet.cell(row=i, column=7)
        cell_calendar_url = excel_sheet.cell(row=i, column=8)
        cell_mfa_enabled = excel_sheet.cell(row=i, column=9)
        json_file['first_name'] = cell_first_name.value
        json_file['last_name'] = cell_last_name.value
        json_file['username'] = cell_username.value
        json_file['email'] = cell_email.value
        json_file['is_active'] = cell_is_active.value
        json_file['nick_name'] = cell_nick_name.value
        json_file['display_name'] = cell_display_name.value
        json_file['calendar_url'] = cell_calendar_url.value
        json_file['mfa_enabled'] = cell_mfa_enabled.value
        response = requests.post('http://badger.boast.net:8080/api/users/user/', json=json_file,
                                 auth=credentials)
        assert response.status_code == 201
        # pdb.set_trace()



